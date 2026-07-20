"""
Servidor de sincronización de productos (Deltron -> Hasbun API).

Corre como un proceso independiente en el VPS (con uvicorn/pm2/systemd/Dokploy).
Aunque cierres el navegador, el proceso sigue trabajando en segundo plano.
Cada vez que quieras ver el avance, entras a la URL del servidor y listo.

Endpoints:
    GET  /                          -> interfaz visual (misma que el HTML anterior)
    POST /api/subir                 -> recibe el Excel y arranca el proceso en background
    GET  /api/estado                -> estado actual (progreso, logs, stats) en JSON
    GET  /api/descargar-fallidos    -> descarga el Excel con los productos que fallaron
"""

import asyncio
import io
import os
import uuid
from datetime import datetime
from urllib.parse import parse_qs, urlparse, quote

import httpx
import openpyxl
from fastapi import FastAPI, File, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse

# ============================================================
# CONFIGURACIÓN (puedes sobreescribir con variables de entorno)
# ============================================================
API_BASE = "https://backend.inversioneshasbun.com/api"
# Nuevo scraper "sego": recibe la URL completa del producto como parámetro `url`
SCRAPER_URL = "https://pc.skrifna.uk/hasbun/scrapers/sego"
STOCK_DEFAULT = int("10")
CONCURRENCIA = int("4")          # 4 cabezas de IA en paralelo, EN LOTES
TIMEOUT_SCRAPER = float("120.0")  # el scraper con IA puede tardar
REINTENTOS_SCRAPER = int("2")  # reintentos ante error de red/timeout
DELAY_ENTRE_REINTENTOS = float("2.0")
UPLOAD_DIR = "/tmp/sincro_uploads"

os.makedirs(UPLOAD_DIR, exist_ok=True)

app = FastAPI(title="Sincronizador de Productos")

# ============================================================
# ESTADO GLOBAL DEL TRABAJO (vive en memoria del proceso)
# ============================================================
def estado_inicial():
    return {
        "activo": False,
        "terminado": False,
        "error": None,
        "archivo": None,
        "titulo": "Esperando archivo",
        "detalle": "",
        "procesados": 0,
        "total": 0,
        "stats": {"categorias": 0, "exitosos": 0, "duplicados": 0, "fallidos": 0},
        "logs": [],
        "reporte_fallidos": [],
    }


JOB = estado_inicial()
CACHE = {"categorias": [], "marcas": [], "codigos_existentes": set()}
MARCA_LOCK = asyncio.Lock()
CATEGORIA_LOCK = asyncio.Lock()


def horaactual():
    return datetime.now().strftime("%H:%M:%S")


def log_add(tipo, titulo, sub=""):
    entrada = {"id": str(uuid.uuid4()), "tipo": tipo, "titulo": titulo, "sub": sub, "hora": horaactual()}
    JOB["logs"].append(entrada)
    return entrada["id"]


def log_update(id_, tipo, titulo=None, sub=None):
    for e in JOB["logs"]:
        if e["id"] == id_:
            e["tipo"] = tipo
            if titulo is not None:
                e["titulo"] = titulo
            if sub is not None:
                e["sub"] = sub
            e["hora"] = horaactual()
            return


# ============================================================
# HELPERS DE RED
# ============================================================
async def fetch_json_safe(client: httpx.AsyncClient, url: str, reintentos: int = REINTENTOS_SCRAPER):
    ultimo_error = None
    for intento in range(reintentos + 1):
        try:
            resp = await client.get(url, timeout=TIMEOUT_SCRAPER)
            try:
                body = resp.json()
            except Exception:
                body = None
            if resp.status_code >= 400:
                if isinstance(body, dict) and body.get("error"):
                    return {"__httpError": True, "status": resp.status_code, "message": body.get("error")}
                return {"__httpError": True, "status": resp.status_code, "message": f"HTTP {resp.status_code}"}
            return body if body is not None else {"__httpError": True, "status": resp.status_code, "message": "Respuesta vacía"}
        except Exception as e:
            ultimo_error = str(e)
            if intento < reintentos:
                await asyncio.sleep(DELAY_ENTRE_REINTENTOS)
                continue
    return {"__networkError": True, "message": ultimo_error}


def safe_json(resp: httpx.Response):
    try:
        return resp.json()
    except Exception:
        return None


def a_numero(valor):
    """Convierte celdas de Excel (que pueden venir como '#VALUE!', None, string, etc.) a float seguro."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace(",", "")
    if not texto or texto.upper() in ("#VALUE!", "#N/A", "#REF!", "#DIV/0!"):
        return None
    try:
        return float(texto)
    except ValueError:
        return None


# ============================================================
# LECTURA DEL EXCEL -> { categoria: [ {codigo, url, precio_compra, precio_venta}, ... ] }
# ============================================================
def extraer_codigo(cell):
    if cell is None:
        return None
    if getattr(cell, "hyperlink", None) is not None and cell.hyperlink.target:
        try:
            parsed = urlparse(cell.hyperlink.target)
            qs = parse_qs(parsed.query)
            for clave in ("item_number", "codee", "code"):
                if clave in qs and qs[clave]:
                    return qs[clave][0].strip().upper()
        except Exception:
            pass
    if cell.value:
        return str(cell.value).strip().upper()
    return None


def leer_excel(path: str):
    """
    Columnas esperadas en cada hoja (fila 1 = encabezado):
        A: Código (con hipervínculo a la URL del producto en Sego)
        B: Precio Compra $
        C: P.Compra s/.
        D: P.Compra Total S/.   <- precio de compra final (se usa como PRECIO_COMPRA)
        E: P.Venta
        F: P.V. FINAL           <- precio de venta final (se usa como PRECIO_VENTA)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    estructura = {}
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        categoria_nombre = nombre_hoja.strip().replace(" ", "-")
        productos = []
        for fila in ws.iter_rows(min_row=2, max_col=6):
            cell_codigo = fila[0]
            codigo = extraer_codigo(cell_codigo)
            if not codigo:
                continue

            url_link = cell_codigo.hyperlink.target if getattr(cell_codigo, "hyperlink", None) else None

            precio_compra = a_numero(fila[3].value) if len(fila) > 3 else None   # columna D
            precio_venta = a_numero(fila[5].value) if len(fila) > 5 else None    # columna F

            productos.append({
                "codigo": codigo,
                "url": url_link,
                "precio_compra": precio_compra,
                "precio_venta": precio_venta,
            })
        if productos:
            estructura[categoria_nombre] = productos
    return estructura


# ============================================================
# CATÁLOGO (categorías / marcas / productos existentes)
# ============================================================
async def cargar_cache_inicial(client: httpx.AsyncClient):
    log_add("info", "Descargando catálogo actual...", "categorías, marcas y productos existentes")

    async def get_lista(url):
        try:
            r = await client.get(url)
            data = safe_json(r)
            return data if isinstance(data, list) else []
        except Exception:
            return []

    categorias, marcas, productos = await asyncio.gather(
        get_lista(f"{API_BASE}/CATEGORIA"),
        get_lista(f"{API_BASE}/MARCA"),
        get_lista(f"{API_BASE}/PRODUCTOS"),
    )

    CACHE["categorias"] = categorias
    CACHE["marcas"] = marcas
    CACHE["codigos_existentes"] = {
        str(p.get("KODIGO_FABIRKANTE")).strip().upper()
        for p in productos
        if p.get("KODIGO_FABIRKANTE")
    }

    log_add(
        "success",
        f"Catálogo cargado: {len(categorias)} categorías, {len(marcas)} marcas, "
        f"{len(CACHE['codigos_existentes'])} productos ya existentes.",
    )


async def asegurar_categoria(client: httpx.AsyncClient, nombre: str):
    nombre_norm = nombre.strip()
    async with CATEGORIA_LOCK:
        existente = next(
            (c for c in CACHE["categorias"] if str(c.get("NOMBRE", "")).lower() == nombre_norm.lower()), None
        )
        if existente:
            log_add("cat", f'Categoría "{nombre_norm}" ya existe (ID {existente["ID"]}) — se sube ahí, no se crea de nuevo.')
            return existente["ID"]

        log_add("cat", f'Creando nueva categoría: "{nombre_norm}"...')
        try:
            resp = await client.post(f"{API_BASE}/CATEGORIA", json={"NOMBRE": nombre_norm})
            lista = safe_json(resp)
        except Exception as e:
            log_add("error", f'No se pudo crear la categoría "{nombre_norm}": {e}')
            return None

        if isinstance(lista, list):
            CACHE["categorias"] = lista

        existente = next(
            (c for c in CACHE["categorias"] if str(c.get("NOMBRE", "")).lower() == nombre_norm.lower()), None
        )
        if existente:
            log_add("success", f'Categoría "{nombre_norm}" creada correctamente (ID {existente["ID"]}).')
            return existente["ID"]

        log_add("error", f'No se pudo crear ni encontrar la categoría "{nombre_norm}".')
        return None


async def asegurar_marca(client: httpx.AsyncClient, nombre: str):
    nombre = (nombre or "GENERICO").strip() or "GENERICO"
    async with MARCA_LOCK:
        existente = next(
            (m for m in CACHE["marcas"] if str(m.get("NOMBRE", "")).lower() == nombre.lower()), None
        )
        if existente:
            return existente["ID"]

        log_add("info", f'Creando nueva marca: "{nombre}"...')
        try:
            resp = await client.post(f"{API_BASE}/MARCA", json={"NOMBRE": nombre})
            lista = safe_json(resp)
        except Exception as e:
            log_add("error", f'No se pudo crear la marca "{nombre}": {e}')
            return None

        if isinstance(lista, list):
            CACHE["marcas"] = lista

        existente = next(
            (m for m in CACHE["marcas"] if str(m.get("NOMBRE", "")).lower() == nombre.lower()), None
        )
        if existente:
            log_add("success", f'Marca "{nombre}" creada correctamente (ID {existente["ID"]}).')
            return existente["ID"]

        log_add("error", f'No se pudo crear ni encontrar la marca "{nombre}".')
        return None


# ============================================================
# VALIDACIÓN + PAYLOAD
# ============================================================
def obtener_codigo_fabrica(g, codigo_excel):
    """
    La API sego ya NO devuelve `general.code`. Ahora manda el SKU en
    `general.kd_fabrica` (y también en `general.modelo`, duplicado).
    Se prueban esos campos en orden y, si ninguno vino, se usa el código
    que ya conocíamos por el propio Excel (columna A) como último respaldo.
    """
    return (
        g.get("code")
        or g.get("kd_fabrica")
        or g.get("modelo")
        or codigo_excel
        or None
    )


def validar_ficha_scraper(json_, precio_compra_excel, precio_venta_excel, codigo_excel):
    if json_.get("__networkError"):
        return [f"No se pudo contactar al scraper (¿está corriendo en {SCRAPER_URL}?): {json_['message']}"]
    if json_.get("__httpError"):
        return [json_.get("message") or f"El scraper respondió con error HTTP {json_.get('status')}"]
    if json_.get("ok") is not True:
        return [json_.get("error") or "El scraper devolvió ok:false"]

    d = json_.get("data") or {}
    g = d.get("general") or {}

    problemas = []
    if not g.get("nombre"):
        problemas.append("Falta el nombre del producto")
    if not obtener_codigo_fabrica(g, codigo_excel):
        problemas.append("Falta el código de fábrica")
    if not d.get("descripcion"):
        problemas.append("Falta la descripción")
    # Los precios ya NO se validan desde el scraper: se toman del Excel.
    if not precio_venta_excel or precio_venta_excel <= 0:
        problemas.append("Falta o es inválido el precio de venta en el Excel")
    if not precio_compra_excel or precio_compra_excel <= 0:
        problemas.append("Falta o es inválido el precio de compra en el Excel")
    if not d.get("imagenes"):
        problemas.append("No tiene imágenes")
    if not g.get("img_principal"):
        problemas.append("Falta la imagen principal")
    return problemas


def construir_payload(json_, categoria_id, marca_id, precio_compra_excel, precio_venta_excel, codigo_excel):
    d = json_["data"]
    g = d["general"]
    urls = d.get("urls_automatizadas") or {}

    return {
        "NOMBRE": g.get("nombre"),
        "KODIGO_FABIRKANTE": obtener_codigo_fabrica(g, codigo_excel),
        "IMG_PROTADA": g.get("img_principal"),
        "DESCRIPCION": d.get("descripcion"),
        "AUDIO": urls.get("audio_descripcion") or "",
        "STOCK": STOCK_DEFAULT,
        "CATEGORIA_ID": categoria_id,
        "MARCA_ID": marca_id,
        "precio": {
            # Precios tomados directamente del Excel (columnas P.Compra Total S/. y P.V. FINAL)
            "PRECIO_COMPRA": precio_compra_excel or 0,
            "PRECIO_VENTA": precio_venta_excel or 0,
            "DESCUENTO_PORCENTAJE": 0,
        },
        "estado": {"ESTADO": "disponible", "VALUE": "Stock disponible"},
        "ficha": {
            # La ficha "compleja/completa" que antes devolvía el scraper pasa a ser la SIMPLE.
            "SIMPLE": (urls.get("ficha_avanzada") or {}).get("url_frontend") or "",
            # La ficha COMPLEJA ahora es el PDF de ficha técnica que devuelve la API.
            "COMPLEJO": d.get("ficha_tecnica_pdf") or urls.get("ficha_tecnica_pdf") or g.get("ficha_tecnica_pdf") or "",
        },
        "componentes": [
            {"icon": c.get("icon"), "titulo": c.get("titulo"), "nombre": c.get("valor")}
            for c in (d.get("mejores_componentes") or [])
        ],
        "imgs": [{"URL": u} for u in (d.get("imagenes") or [])],
        "tags": d.get("etiquetas") or [],
    }


# ============================================================
# PIPELINE PRINCIPAL (corre en background, sobrevive al cierre del navegador)
# ============================================================
async def procesar_archivo(path: str, nombre_archivo: str):
    JOB.update(
        activo=True,
        terminado=False,
        error=None,
        archivo=nombre_archivo,
        titulo="Leyendo Archivo",
        detalle=f'Analizando "{nombre_archivo}"...',
        procesados=0,
        total=0,
        logs=[],
        reporte_fallidos=[],
        stats={"categorias": 0, "exitosos": 0, "duplicados": 0, "fallidos": 0},
    )

    try:
        log_add("info", f"Archivo recibido: {nombre_archivo}")
        estructura = leer_excel(path)

        if not estructura:
            log_add("error", "No se encontraron productos con enlaces/códigos válidos en el archivo.")
            JOB["titulo"] = "Sin Datos"
            JOB["detalle"] = "El archivo no contiene productos reconocibles."
            return

        log_add("success", f"Se detectaron {len(estructura)} categorías en el archivo.")

        limites = httpx.Limits(max_connections=CONCURRENCIA + 2, max_keepalive_connections=CONCURRENCIA)
        timeout_cliente = httpx.Timeout(TIMEOUT_SCRAPER, connect=30.0)
        async with httpx.AsyncClient(timeout=timeout_cliente, limits=limites) as client:
            await cargar_cache_inicial(client)

            total = sum(len(v) for v in estructura.values())
            JOB["total"] = total
            procesados_en_run = set()

            # --- 1. Asegurar categorías (secuencial, son pocas) ---
            categoria_ids = {}
            for nombre_categoria, items in estructura.items():
                JOB["titulo"] = "Categorizando"
                JOB["detalle"] = f"Preparando categoría: {nombre_categoria}"
                cid = await asegurar_categoria(client, nombre_categoria)
                JOB["stats"]["categorias"] += 1
                categoria_ids[nombre_categoria] = cid

                if cid is None:
                    log_add("error", f'Se omite toda la categoría "{nombre_categoria}" por no poder crearse.')
                    for item in items:
                        JOB["stats"]["fallidos"] += 1
                        JOB["reporte_fallidos"].append({
                            "Codigo": item["codigo"], "Categoria": nombre_categoria,
                            "Motivo": "No se pudo crear/encontrar la categoría", "URL": item.get("url") or "",
                        })
                        JOB["procesados"] += 1

            # --- 2. Lista plana de productos a procesar (de categorías válidas) ---
            tareas = []
            for nombre_categoria, items in estructura.items():
                cid = categoria_ids[nombre_categoria]
                if cid is None:
                    continue
                for item in items:
                    tareas.append((nombre_categoria, cid, item))

            async def procesar_producto(nombre_categoria, categoria_id, item):
                    codigo = item["codigo"]
                    url_producto = item.get("url")
                    precio_compra_excel = item.get("precio_compra")
                    precio_venta_excel = item.get("precio_venta")

                    JOB["titulo"] = "Procesando Producto"
                    JOB["detalle"] = f"{nombre_categoria} → {codigo}"

                    if codigo in procesados_en_run:
                        log_add("warn", f"{codigo}: repetido dentro del mismo archivo, se omite.")
                        JOB["stats"]["duplicados"] += 1
                        JOB["procesados"] += 1
                        return
                    procesados_en_run.add(codigo)

                    if codigo in CACHE["codigos_existentes"]:
                        log_add("warn", f"{codigo}: ya existe en el sistema, no se vuelve a subir.")
                        JOB["stats"]["duplicados"] += 1
                        JOB["reporte_fallidos"].append({
                            "Codigo": codigo, "Categoria": nombre_categoria,
                            "Motivo": "Duplicado: ya existía en el sistema", "URL": url_producto or "",
                        })
                        JOB["procesados"] += 1
                        return

                    if not url_producto:
                        log_add("error", f"{codigo}: no tiene URL de producto (hipervínculo) en el Excel, se omite.")
                        JOB["stats"]["fallidos"] += 1
                        JOB["reporte_fallidos"].append({
                            "Codigo": codigo, "Categoria": nombre_categoria,
                            "Motivo": "Sin URL de producto en el Excel", "URL": "",
                        })
                        JOB["procesados"] += 1
                        return

                    log_id = log_add("loading", codigo, f"{nombre_categoria} · consultando scraper...")

                    scraper_json = await fetch_json_safe(
                        client, f"{SCRAPER_URL}?url={quote(url_producto, safe='')}"
                    )
                    problemas = validar_ficha_scraper(scraper_json, precio_compra_excel, precio_venta_excel, codigo)

                    if problemas:
                        log_update(log_id, "error", codigo, " | ".join(problemas))
                        JOB["stats"]["fallidos"] += 1
                        JOB["reporte_fallidos"].append({
                            "Codigo": codigo, "Categoria": nombre_categoria,
                            "Motivo": " | ".join(problemas), "URL": url_producto or "",
                        })
                        JOB["procesados"] += 1
                        return

                    marca_nombre = (scraper_json.get("data", {}).get("general", {}) or {}).get("marca")
                    marca_id = await asegurar_marca(client, marca_nombre)
                    payload = construir_payload(
                        scraper_json, categoria_id, marca_id, precio_compra_excel, precio_venta_excel, codigo
                    )

                    try:
                        resp = await client.post(f"{API_BASE}/PRODUCTOS", json=payload)
                        body = safe_json(resp)
                        if resp.status_code >= 400:
                            mensaje = body.get("error") if isinstance(body, dict) else f"HTTP {resp.status_code}"
                            raise RuntimeError(mensaje)

                        log_update(log_id, "success", payload["NOMBRE"], f'{codigo} · subido correctamente a "{nombre_categoria}"')
                        JOB["stats"]["exitosos"] += 1
                        CACHE["codigos_existentes"].add(codigo)
                    except Exception as e:
                        log_update(log_id, "error", codigo, f"Error al subir a la API: {e}")
                        JOB["stats"]["fallidos"] += 1
                        JOB["reporte_fallidos"].append({
                            "Codigo": codigo, "Categoria": nombre_categoria,
                            "Motivo": f"Error al subir a la API: {e}", "URL": url_producto or "",
                        })

                    JOB["procesados"] += 1

            # --- 3. Procesar en LOTES ESTRICTOS de CONCURRENCIA productos.
            #        Se lanzan los N del lote, se espera a que TODOS terminen
            #        (éxito o fallo) y recién ahí se lanza el siguiente lote.
            #        Nada de tareas sueltas circulando de fondo: así se evita
            #        que un producto lento arrastre al resto y el cliente
            #        HTTP termine en un estado inconsistente ("client closed").
            for inicio in range(0, len(tareas), CONCURRENCIA):
                lote = tareas[inicio:inicio + CONCURRENCIA]
                JOB["detalle"] = f"Lote {inicio // CONCURRENCIA + 1} de {(len(tareas) - 1) // CONCURRENCIA + 1} ({len(lote)} productos)"
                tasks = [
                    asyncio.create_task(procesar_producto(nombre_categoria, cid, item))
                    for (nombre_categoria, cid, item) in lote
                ]
                await asyncio.gather(*tasks, return_exceptions=True)

        JOB["titulo"] = "Finalizado"
        JOB["detalle"] = "Sincronización completa."
        log_add("info", "Proceso completado. Resumen listo.")

    except Exception as e:
        log_add("error", f"Error inesperado del sistema: {e}")
        JOB["error"] = str(e)
    finally:
        JOB["activo"] = False
        JOB["terminado"] = True
        try:
            os.remove(path)
        except OSError:
            pass


# ============================================================
# ENDPOINTS
# ============================================================
@app.get("/", response_class=HTMLResponse)
async def index():
    ruta_html = os.path.join(os.path.dirname(__file__), "index.html")
    with open(ruta_html, "r", encoding="utf-8") as f:
        return f.read()


@app.post("/api/subir")
async def subir(file: UploadFile = File(...)):
    if JOB["activo"]:
        return JSONResponse({"ok": False, "error": "Ya hay un proceso en curso. Espera a que termine."}, status_code=409)

    nombre_seguro = f"{uuid.uuid4()}_{file.filename}"
    ruta = os.path.join(UPLOAD_DIR, nombre_seguro)
    contenido = await file.read()
    with open(ruta, "wb") as f:
        f.write(contenido)

    asyncio.create_task(procesar_archivo(ruta, file.filename))
    return {"ok": True}


@app.get("/api/estado")
async def estado():
    return JOB


@app.get("/api/descargar-fallidos")
async def descargar_fallidos():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fallidos"
    ws.append(["Codigo", "Categoria", "Motivo", "URL"])
    for fila in JOB["reporte_fallidos"]:
        ws.append([fila.get("Codigo", ""), fila.get("Categoria", ""), fila.get("Motivo", ""), fila.get("URL", "")])

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 50

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha = datetime.now().strftime("%Y-%m-%d")
    headers = {"Content-Disposition": f'attachment; filename="productos_fallidos_{fecha}.xlsx"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=int(os.environ.get("PORT", "8000")), reload=False)
